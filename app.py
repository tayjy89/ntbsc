#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
NTBSC Roster Planner — single-file Streamlit app (Python 3.11+)
Features:
- Password-gated admin tabs (public sees Overview + Exports only)
- Editors for consultants, Room28 defaults, duties (single CSV with kinds), juniors, paeds, specials
- Duty entry by date range (validated dates; dropdowns for kind/session)
- Multi-month roster building/persistence (data/rosters/YYYY-MM.csv)
- Weekly-smart preceptor selection (fixed mode) with smoothing
- Suma auto-assign on 2nd & 4th Friday PM
- Blocking: PH, leave, woodlands, ward (AM), blues (PM)
- Conflicts detection + highlighting
- Editable roster grid
- Overview with consultant filter
- Calendar view with 🎉 icon and shaded PH cells; month/year toggle affects overview list too
- Exports: CSV & Excel (Roster, Overview_List, Export_Flat)
- Singapore public holidays hard-wired (2024–2026)
- NaN-safe normalisation to avoid TypeError on joins/exports
"""

import calendar
import json
from io import BytesIO
from pathlib import Path
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "NTBSC Roster Planner"
DATA_DIR = Path("data")
ROSTER_DIR = DATA_DIR / "rosters"
DATA_DIR.mkdir(exist_ok=True, parents=True)
ROSTER_DIR.mkdir(exist_ok=True, parents=True)

# Password is kept in code for app usage; do not publish it elsewhere.
ADMIN_PASSWORD = "NTBSC_jy"

# -------------------- Public Holidays (SG) --------------------
def sg_public_holidays(year: int) -> set[date]:
    fixed = {
        2024: [
            "2024-01-01","2024-02-10","2024-02-11","2024-03-29","2024-04-10",
            "2024-05-01","2024-05-22","2024-06-17","2024-08-09","2024-10-31","2024-12-25"
        ],
        2025: [
            "2025-01-01","2025-01-29","2025-01-30","2025-04-18","2025-05-01",
            "2025-05-12","2025-06-06","2025-08-09","2025-10-20","2025-12-25"
        ],
        2026: [
            "2026-01-01","2026-02-17","2026-02-18","2026-04-03","2026-05-01",
            "2026-05-22","2026-06-26","2026-08-09","2026-11-09","2026-12-25"
        ],
    }
    return set(pd.to_datetime(fixed.get(year, []), errors="coerce").dropna().date)

# -------------------- IO helpers --------------------
def ensure_seed_files():
    cfp = DATA_DIR / "consultants.csv"
    if not cfp.exists():
        pd.DataFrame([
            {"name": "Jun Yang", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "Y"},
            {"name": "Deborah", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "Y"},
            {"name": "Wilnard", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "Y"},
            {"name": "Suma", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "N"},
            {"name": "Matthias", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "N"},
            {"name": "Khin", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "Y"},
            {"name": "Hoi Wah", "active(Y/N)": "Y", "preceptor_eligible(Y/N)": "N"},
        ]).to_csv(cfp, index=False)

    dfp = DATA_DIR / "room28_defaults.csv"
    if not dfp.exists():
        pd.DataFrame([
            {"weekday": "Mon", "session": "AM", "consultant": "Khin"},
            {"weekday": "Mon", "session": "PM", "consultant": "Jun Yang"},
            {"weekday": "Tue", "session": "AM", "consultant": "Matthias"},
            {"weekday": "Tue", "session": "PM", "consultant": "Wilnard"},
            {"weekday": "Wed", "session": "AM", "consultant": "Hoi Wah"},
            {"weekday": "Wed", "session": "PM", "consultant": "Deborah"},
            {"weekday": "Thu", "session": "AM", "consultant": ""},
            {"weekday": "Thu", "session": "PM", "consultant": "Khin"},
            {"weekday": "Fri", "session": "AM", "consultant": "Hoi Wah"},
            {"weekday": "Fri", "session": "PM", "consultant": "Khin"},
        ]).to_csv(dfp, index=False)

    sfp = DATA_DIR / "settings.json"
    if not sfp.exists():
        today = date.today()
        json.dump({"month": today.month, "year": today.year, "week_preceptor_mode": "fixed"}, open(sfp, "w"))

    # Single duties file (kinds: leave/ward/blues/woodlands/others)
    base_csvs = [
        ("duties.csv", ["consultant","date","session","kind","notes"]),
        ("juniors.csv", ["date","junior_name"]),
        ("paeds.csv", ["consultant","date","session","notes"]),
        ("special_slots.csv", ["date","session","room_name","assignee"]),
    ]
    for fname, cols in base_csvs:
        fp = DATA_DIR / fname
        if not fp.exists():
            pd.DataFrame(columns=cols).to_csv(fp, index=False)

def read_csv(name: str, parse_dates=("date",)) -> pd.DataFrame:
    fp = DATA_DIR / name
    if not fp.exists():
        pd.DataFrame().to_csv(fp, index=False)
    df = pd.read_csv(fp, dtype=str).fillna("")
    for col in parse_dates:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
    return df

def write_csv(name: str, df: pd.DataFrame) -> None:
    df2 = df.copy()

    # Replace NaN/None with "" in non-date columns; normalise "nan" strings
    for c in df2.columns:
        if c.lower() != "date":
            df2[c] = df2[c].where(~pd.isna(df2[c]), "")
            df2[c] = df2[c].map(lambda x: "" if pd.isna(x) else ("" if str(x).strip().lower()=="nan" else str(x)))

    def _norm(x):
        if hasattr(x, "isoformat"):  # dates
            return x.isoformat()
        if pd.isna(x) or x is None:
            return ""
        sx = str(x)
        return "" if sx.strip().lower()=="nan" else sx

    for c in df2.columns:
        df2[c] = df2[c].map(_norm)

    df2.to_csv(DATA_DIR / name, index=False)

def read_settings() -> dict: 
    return json.load(open(DATA_DIR / "settings.json"))

def write_settings(s: dict) -> None: 
    json.dump(s, open(DATA_DIR / "settings.json","w"))

# -------------------- Date helpers --------------------
def weekdays_in_month(year: int, month: int):
    cal = calendar.Calendar(firstweekday=calendar.MONDAY)
    out = []
    for d in cal.itermonthdates(year, month):
        if d.month == month and d.weekday() < 5:
            out.append((d, ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d.weekday()]))
    return out

def second_and_fourth_fridays(year: int, month: int):
    fridays = [d for d, wd in weekdays_in_month(year, month) if wd == "Fri"]
    out = []
    if len(fridays) >= 2: out.append(fridays[1])
    if len(fridays) >= 4: out.append(fridays[3])
    return out

# -------------------- Lookups / Blocks --------------------
def _room28_default_for(wd, session, defaults_df):
    m = defaults_df[(defaults_df["weekday"]==wd) & (defaults_df["session"]==session)]
    return m.iloc[0]["consultant"].strip() if not m.empty else ""

def _is_ph(d: date, year_ph_set: set) -> bool: 
    return d in year_ph_set

def _has_leave(consultant, d, session, duties_df) -> bool:
    rows = duties_df[(duties_df["consultant"]==consultant) & (duties_df["date"]==d) & (duties_df["kind"]=="leave")]
    for _, r in rows.iterrows():
        s = str(r.get("session","")).upper()
        if s in ("FULL","FD","WHOLE","ALL","DAY","FULL DAY",""): return True
        if s == session.upper(): return True
    return False

def _has_woodlands(consultant, d, duties_df) -> bool:
    return not duties_df[(duties_df["consultant"]==consultant) & (duties_df["date"]==d) & (duties_df["kind"]=="woodlands")].empty

def _has_ward_am(consultant, d, duties_df) -> bool:
    return not duties_df[(duties_df["consultant"]==consultant) & (duties_df["date"]==d) & (duties_df["kind"]=="ward")].empty

def _has_blues_pm(consultant, d, duties_df) -> bool:
    return not duties_df[(duties_df["consultant"]==consultant) & (duties_df["date"]==d) & (duties_df["kind"]=="blues")].empty

# -------------------- Build roster --------------------
def build_roster(year, month, week_preceptor_mode, consultants_df, defaults_df, juniors_df, year_ph_set, duties_df, paeds_df, special_df) -> pd.DataFrame:
    preceptor_eligible = set(
        consultants_df[(consultants_df["active(Y/N)"]=="Y") & (consultants_df["preceptor_eligible(Y/N)"]=="Y")]["name"].tolist()
    )
    day_slots = weekdays_in_month(year, month)
    juniors_map = {r["date"]: r["junior_name"] for _, r in juniors_df.iterrows() if r["date"] is not None and r["date"]!=""}
    fri_2_4 = second_and_fourth_fridays(year, month)

    # Weekly preceptor selection (fixed mode) with smoothing & availability
    week_monday_to_preceptor = {}
    if week_preceptor_mode == "fixed":
        def _available(cand, d, sess):
            if _is_ph(d, year_ph_set): return False
            wd = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d.weekday()]
            r28 = _room28_default_for(wd, sess, defaults_df)
            if wd=="Fri" and sess=="PM" and d in fri_2_4: r28 = "Suma"
            if r28 == cand: return False
            if _has_woodlands(cand, d, duties_df): return False
            if _has_leave(cand, d, sess, duties_df): return False
            if sess=="AM" and _has_ward_am(cand, d, duties_df): return False
            if sess=="PM" and _has_blues_pm(cand, d, duties_df): return False
            return True

        dates = [d for d, _ in day_slots]
        if dates:
            start = min(dates); start_m = start - timedelta(days=start.weekday())
            end = max(dates); end_su = end + timedelta(days=(6-end.weekday()))
            prior = {c:0 for c in preceptor_eligible}
            cur = start_m
            while cur <= end_su:
                wk = [cur+timedelta(days=i) for i in range(5) if (cur+timedelta(days=i)).month==month]
                best, best_score = None, -1
                for cand in preceptor_eligible:
                    score = sum(_available(cand, d, s) for d in wk for s in ("AM","PM"))
                    if score>best_score or (score==best_score and best is not None and prior[cand]<prior[best]) or (score==best_score and prior[cand]==prior[best] and cand<best):
                        if score>=0: best, best_score = cand, score
                week_monday_to_preceptor[cur] = best if best_score>0 else None
                if best and best_score>0: prior[best]+=1
                cur += timedelta(days=7)

    rows = []
    for d, wd in day_slots:
        for session in ("AM","PM"):
            room18 = juniors_map.get(d, "TBD")
            room28 = _room28_default_for(wd, session, defaults_df)
            if wd=="Fri" and session=="PM" and d in fri_2_4:
                room28 = "Suma"
            room29 = ""
            for _, r in paeds_df[paeds_df["date"]==d].iterrows():
                if str(r.get("session","")).upper() in ("", session.upper()):
                    room29 = r["consultant"]
            notes = []
            specials_today = special_df[(special_df["date"]==d) & (special_df["session"].str.upper().isin(["", session.upper()]))] if not special_df.empty else pd.DataFrame()
            for _, r in specials_today.iterrows():
                rn, asg = r.get("room_name",""), r.get("assignee","")
                if str(rn).strip().lower() in ("room29","room 29","r29","29"):
                    if not room29: room29 = asg
                elif rn or asg:
                    notes.append(f"{rn}: {asg}")
            preceptor = ""
            if week_preceptor_mode=="fixed":
                monday = d - timedelta(days=d.weekday())
                cand = week_monday_to_preceptor.get(monday, "")
                preceptor = cand or ""
            rows.append({"date":d,"weekday":wd,"session":session,"Room 18":room18,"Room 28":room28,"Preceptor":preceptor,"Room 29":room29,"notes":"; ".join(notes) if notes else ""})

    roster = pd.DataFrame(rows)

    # Apply blocking
    def _wipe(i, cols):
        for c in cols: roster.loc[i, c] = ""

    for i in roster.index:
        r = roster.loc[i]; d, session = r["date"], r["session"]
        if _is_ph(d, sg_public_holidays(d.year)):
            _wipe(i, ["Room 18","Room 28","Preceptor","Room 29"])
            roster.loc[i, "notes"] = (str(roster.loc[i, "notes"]) + "; Public Holiday").strip("; ")
            continue
        for col in ("Room 28","Preceptor"):
            nm = roster.loc[i, col]
            if not nm: continue
            if _has_woodlands(nm, d, duties_df):
                roster.loc[i, col] = ""; roster.loc[i,"notes"] = (str(roster.loc[i,"notes"])+f"; Woodlands({nm})").strip("; ")
            elif _has_leave(nm, d, session, duties_df):
                roster.loc[i, col] = ""; roster.loc[i,"notes"] = (str(roster.loc[i,"notes"])+f"; Leave({nm})").strip("; ")
        for col in ("Room 28","Preceptor"):
            nm = roster.loc[i, col]
            if nm and session=="AM" and _has_ward_am(nm, d, duties_df):
                roster.loc[i, col] = ""; roster.loc[i,"notes"] = (str(roster.loc[i,"notes"])+f"; WardRounds({nm})").strip("; ")
            if nm and session=="PM" and _has_blues_pm(nm, d, duties_df):
                roster.loc[i, col] = ""; roster.loc[i,"notes"] = (str(roster.loc[i,"notes"])+f"; Blues({nm})").strip("; ")
        if roster.loc[i,"Room 28"] and roster.loc[i,"Preceptor"] and roster.loc[i,"Room 28"]==roster.loc[i,"Preceptor"]:
            roster.loc[i,"Preceptor"] = ""
    return roster

# -------------------- Conflicts & styling --------------------
def conflicts(roster_df, consultants_df, year_ph_set, duties_df) -> pd.DataFrame:
    rows = []
    pre_ok = set(consultants_df[(consultants_df["active(Y/N)"]=="Y") & (consultants_df["preceptor_eligible(Y/N)"]=="Y")]["name"].tolist())

    def _add(idx, d, s, field, name, reason):
        rows.append({"row_index": idx, "date": d, "session": s, "field": field, "name": name, "conflict_reason": reason})

    for idx, r in roster_df.reset_index().iterrows():
        d, s = r["date"], r["session"]
        if d in year_ph_set:
            for f in ["Room 18","Room 28","Preceptor","Room 29"]:
                if str(r[f]).strip():
                    _add(r["index"], d, s, f, r[f], "Scheduled on public holiday")
        for f in ["Room 28","Preceptor"]:
            nm = str(r[f]).strip()
            if not nm: continue
            # leave
            for _, row in duties_df.iterrows():
                if row.get("consultant")==nm and row.get("date")==d and row.get("kind")=="leave":
                    sess = str(row.get("session","")).upper()
                    if sess in ("","FULL","FD") or sess==s:
                        _add(r["index"], d, s, f, nm, "On leave")
                        break
            # woodlands
            if any((row.get("consultant")==nm and row.get("date")==d and row.get("kind")=="woodlands") for _, row in duties_df.iterrows()):
                _add(r["index"], d, s, f, nm, "Woodlands day")
            # ward/blues
            if s=="AM" and any((row.get("consultant")==nm and row.get("date")==d and row.get("kind")=="ward") for _, row in duties_df.iterrows()):
                _add(r["index"], d, s, f, nm, "Ward rounds (AM)")
            if s=="PM" and any((row.get("consultant")==nm and row.get("date")==d and row.get("kind")=="blues") for _, row in duties_df.iterrows()):
                _add(r["index"], d, s, f, nm, "Blues (PM)")
        # eligibility
        nm = str(r["Preceptor"]).strip()
        if nm and nm not in pre_ok:
            _add(r["index"], d, s, "Preceptor", nm, "Not preceptor-eligible")
        # exclusivity
        if str(r["Room 28"]).strip() and str(r["Preceptor"]).strip() and r["Room 28"]==r["Preceptor"]:
            _add(r["index"], d, s, "Preceptor", r["Preceptor"], "Same as Room 28 (exclusivity)")
        if str(r["Room 18"]).strip() and str(r["Preceptor"]).strip() and r["Room 18"]==r["Preceptor"]:
            _add(r["index"], d, s, "Preceptor", r["Preceptor"], "Same as Room 18 (exclusivity)")
    return pd.DataFrame(rows)

def style_roster(roster_df: pd.DataFrame, conf_df: pd.DataFrame):
    styles = pd.DataFrame("", index=roster_df.index, columns=roster_df.columns)
    cmap = {(r["date"], r["session"], r["field"]): True for _, r in conf_df.iterrows()}
    for i, r in roster_df.iterrows():
        for col in ["Room 18","Room 28","Preceptor","Room 29"]:
            if (r["date"], r["session"], col) in cmap:
                styles.at[i, col] = "background-color:#ffe6e6;border:1px solid #ff4d4f;"
            elif (col=="Room 18" and r[col]=="TBD") or (col!="Room 18" and r[col]==""):
                styles.at[i, col] = "background-color:#fff7e6;"
    return roster_df.style.set_properties(**{"font-size":"12px","padding":"4px 6px"}).apply(lambda _: styles, axis=None)

# -------------------- Overview (NaN-safe) --------------------
def overview_list(roster_df: pd.DataFrame) -> pd.DataFrame:
    if roster_df.empty:
        return pd.DataFrame(columns=["date","weekday","Room18_AM","Room28_AM","Preceptor_AM","Room29_AM",
                                     "Room18_PM","Room28_PM","Preceptor_PM","Room29_PM","notes"])

    def _s(x):
        if pd.isna(x) or x is None: return ""
        sx = str(x)
        return "" if sx.strip().lower()=="nan" else sx

    def _compact(day_df):
        day_df = day_df.copy()
        # normalise columns to strings
        for c in ["Room 18","Room 28","Preceptor","Room 29","notes","session","weekday"]:
            if c in day_df.columns:
                day_df[c] = day_df[c].map(_s)
        rec = {"date": day_df.iloc[0]["date"], "weekday": _s(day_df.iloc[0]["weekday"])}
        for _, r in day_df.iterrows():
            sfx = r["session"]
            rec[f"Room18_{sfx}"] = _s(r.get("Room 18"))
            rec[f"Room28_{sfx}"] = _s(r.get("Room 28"))
            rec[f"Preceptor_{sfx}"] = _s(r.get("Preceptor"))
            rec[f"Room29_{sfx}"] = _s(r.get("Room 29"))
        notes_clean = [_s(x) for x in day_df["notes"].tolist() if _s(x)]
        rec["notes"] = "; ".join(notes_clean)
        return rec

    out = []
    for d, grp in roster_df.groupby("date", sort=True):
        out.append(_compact(grp.sort_values("session")))
    return pd.DataFrame(out).sort_values("date")

# -------------------- Calendar --------------------
def calendar_markdown(roster_df: pd.DataFrame, year_ph_set: set):
    if roster_df.empty:
        return "_No roster_"
    year = roster_df["date"].iloc[0].year
    month = roster_df["date"].iloc[0].month
    cal = calendar.Calendar(firstweekday=calendar.MONDAY)
    weeks = cal.monthdatescalendar(year, month)
    idx = {(r["date"], r["session"]): r for _, r in roster_df.iterrows()}

    def cell_html(d):
        am = idx.get((d, "AM")); pm = idx.get((d, "PM"))
        ph = d in year_ph_set
        icon = " 🎉" if ph else ""
        bg = " style='background:#ffecec;border-radius:6px;padding:6px;'" if ph else ""
        def fmt(slot):
            if slot is None: return ""
            parts = []
            r28 = slot.get("Room 28", "")
            p = slot.get("Preceptor", "")
            r29 = slot.get("Room 29", "")
            if r28: parts.append(f"R28: {r28}")
            if p: parts.append(f"P: {p}")
            if r29: parts.append(f"R29: {r29}")
            return "; ".join(parts)
        return f"<div{bg}><strong>{d.day}</strong>{icon}<br/>AM: {fmt(am)}<br/>PM: {fmt(pm)}</div>"

    md = []
    md.append("| Mon | Tue | Wed | Thu | Fri | Sat | Sun |")
    md.append("|-----|-----|-----|-----|-----|-----|-----|")
    for wk in weeks:
        cells = []
        for d in wk:
            if d.month != month:
                cells.append(" ")
            else:
                cells.append(cell_html(d))
        md.append("| " + " | ".join(cells) + " |")
    return "\n".join(md)

# -------------------- Exports --------------------
def excel_bytes(roster_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Roster"
    for r in dataframe_to_rows(roster_df, index=False, header=True): ws1.append(r)
    ws2 = wb.create_sheet("Overview_List")
    ov = overview_list(roster_df)
    for r in dataframe_to_rows(ov, index=False, header=True): ws2.append(r)
    ws3 = wb.create_sheet("Export_Flat")
    for r in dataframe_to_rows(roster_df, index=False, header=True): ws3.append(r)
    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

# -------------------- Normalisation helper --------------------
def _normalise_roster_inplace(df: pd.DataFrame) -> None:
    """Coerce key columns to strings; replace NaN/None/'nan' with '' to avoid join/type errors."""
    for col in ["Room 18","Room 28","Preceptor","Room 29","notes","session","weekday"]:
        if col in df.columns:
            df[col] = df[col].astype("object")
            df[col] = df[col].where(~pd.isna(df[col]), "")
            df[col] = df[col].map(lambda x: "" if pd.isna(x) else ("" if str(x).strip().lower()=="nan" else str(x)))

# -------------------- UI --------------------
def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    ensure_seed_files()
    settings = read_settings()

    # Sidebar settings
    with st.sidebar:
        st.header("Settings")
        col1, col2 = st.columns(2)
        month = col1.selectbox("Month", list(range(1,13)),
                               index=settings.get("month", date.today().month)-1, key="month")
        year = col2.number_input("Year", min_value=2020, max_value=2100,
                                 value=settings.get("year", date.today().year), step=1, key="year")
        mode = st.radio("Preceptor assignment", ["fixed","per_session"],
                        index=0 if settings.get("week_preceptor_mode","fixed")=="fixed" else 1,
                        key="mode")
        pw = st.text_input("Admin password", type="password")
        unlocked = (pw == ADMIN_PASSWORD)
        if st.button("Save Settings"):
            write_settings({"month": int(month), "year": int(year), "week_preceptor_mode": mode})
            st.success("Settings saved.")

    # Load base inputs
    consultants_df = read_csv("consultants.csv", parse_dates=())
    defaults_df = read_csv("room28_defaults.csv", parse_dates=())
    juniors_df = read_csv("juniors.csv")
    duties_df = read_csv("duties.csv")
    paeds_df = read_csv("paeds.csv")
    special_df = read_csv("special_slots.csv")
    year_ph = sg_public_holidays(int(year))

    # Load or build roster for selected month
    roster_fp = ROSTER_DIR / f"{int(year)}-{int(month):02d}.csv"
    if roster_fp.exists():
        roster_df = pd.read_csv(roster_fp, dtype=str)
        roster_df["date"] = pd.to_datetime(roster_df["date"]).dt.date
    else:
        roster_df = build_roster(int(year), int(month), mode,
                                 consultants_df, defaults_df, juniors_df,
                                 year_ph, duties_df, paeds_df, special_df)

    # Normalise roster (NaN-safe)
    _normalise_roster_inplace(roster_df)

    # Tabs (gated)
    if unlocked:
        tabs = st.tabs(["Consultants & Defaults","Duties (All-in-one)","Build / Edit Roster","Overview","Exports"])
    else:
        tabs = st.tabs(["Overview","Exports"])

    # Admin tabs
    if unlocked:
        with tabs[0]:
            st.subheader("Consultants")
            cedit = st.data_editor(
                consultants_df,
                num_rows="dynamic",
                width="stretch",
                key="edit_consultants"
            )
            if st.button("Save Consultants"):
                write_csv("consultants.csv", cedit)
                st.success("Saved consultants.csv")

            st.divider()
            st.subheader("Room 28 Defaults")
            ddef = st.data_editor(
                defaults_df,
                num_rows="dynamic",
                width="stretch",
                key="edit_room28_defaults"
            )
            if st.button("Save Room28 Defaults"):
                write_csv("room28_defaults.csv", ddef)
                st.success("Saved room28_defaults.csv")

        with tabs[1]:
            st.subheader("Add duty by date range")
            with st.form("duty_range"):
                c = st.selectbox("Consultant", consultants_df["name"], key="duties_consultant")
                kind = st.selectbox("Kind", ["leave","ward","blues","woodlands","others"], key="duties_kind")
                sess = st.selectbox("Session", ["AM","PM","Full"], key="duties_session")
                colr1, colr2 = st.columns(2)
                start = colr1.date_input("Start date", key="duties_start")
                end = colr2.date_input("End date", key="duties_end")
                note = st.text_input("Notes", key="duties_notes")
                submitted = st.form_submit_button("Add range")
                if submitted and start and end and start <= end:
                    dr = pd.date_range(start, end, freq="D")
                    new = pd.DataFrame([{"consultant": c, "date": d.date(), "session": sess, "kind": kind, "notes": note} for d in dr])
                    duties_df2 = pd.concat([duties_df, new], ignore_index=True)
                    write_csv("duties.csv", duties_df2)
                    st.success(f"Added {len(new)} rows to duties.csv")
                    duties_df = duties_df2

            st.subheader("Duties editor")
            ded = st.data_editor(
                duties_df,
                column_config={
                    "date": st.column_config.DateColumn("Date"),
                    "kind": st.column_config.SelectboxColumn("Kind", options=["leave","ward","blues","woodlands","others"]),
                    "session": st.column_config.SelectboxColumn("Session", options=["AM","PM","Full"]),
                },
                num_rows="dynamic",
                width="stretch",
                key="duties_editor"
            )
            if st.button("Save Duties"):
                write_csv("duties.csv", ded)
                st.success("Saved duties.csv")

            st.divider()
            st.subheader("Juniors / Paeds / Specials")
            jedit = st.data_editor(read_csv("juniors.csv"), column_config={"date": st.column_config.DateColumn("Date")}, num_rows="dynamic", key="edit_juniors")
            if st.button("Save Juniors"): write_csv("juniors.csv", jedit); st.success("Saved juniors.csv")
            pedit = st.data_editor(read_csv("paeds.csv"), column_config={"date": st.column_config.DateColumn("Date"), "session": st.column_config.SelectboxColumn("Session", options=["AM","PM"])}, num_rows="dynamic", key="edit_paeds")
            if st.button("Save Paeds"): write_csv("paeds.csv", pedit); st.success("Saved paeds.csv")
            sedit = st.data_editor(read_csv("special_slots.csv"), column_config={"date": st.column_config.DateColumn("Date"), "session": st.column_config.SelectboxColumn("Session", options=["AM","PM",""])}, num_rows="dynamic", key="edit_specials")
            if st.button("Save Specials"): write_csv("special_slots.csv", sedit); st.success("Saved special_slots.csv")

        with tabs[2]:
            st.subheader("Build / Edit Roster")
            if st.button("Rebuild from inputs for this month"):
                roster_df = build_roster(int(year), int(month), mode, read_csv("consultants.csv",parse_dates=()), read_csv("room28_defaults.csv",parse_dates=()), read_csv("juniors.csv"), sg_public_holidays(int(year)), read_csv("duties.csv"), read_csv("paeds.csv"), read_csv("special_slots.csv"))
                _normalise_roster_inplace(roster_df)
                st.success("Roster regenerated. You can edit below and save.")
            conf_df = conflicts(roster_df, read_csv("consultants.csv",parse_dates=()), sg_public_holidays(int(year)), read_csv("duties.csv"))
            st.markdown("**Roster (highlighted)**")
            styled = style_roster(roster_df, conf_df)
            st.markdown(styled.to_html(), unsafe_allow_html=True)

            st.markdown("**Edit roster (inline)**")
            redit = st.data_editor(
                roster_df,
                column_config={"date": st.column_config.DateColumn("Date"), "session": st.column_config.SelectboxColumn("Session", options=["AM","PM"])},
                num_rows="dynamic", width="stretch", key="edit_roster"
            )
            if st.button("Save Roster for this Month"):
                df2 = redit.copy(); df2["date"] = pd.to_datetime(df2["date"]).dt.date
                df2.to_csv(ROSTER_DIR / f"{int(year)}-{int(month):02d}.csv", index=False)
                st.success(f"Saved to rosters/{int(year)}-{int(month):02d}.csv")

            st.markdown("**Conflicts**")
            st.dataframe(conf_df, use_container_width=True, height=240)

    # Public Overview
    with tabs[-2 if unlocked else 0]:
        st.subheader("Overview")
        colov1, colov2 = st.columns(2)
        ov_month = colov1.selectbox("Month (Overview/Calendar)", list(range(1,13)),
                                    index=int(month)-1, key="ov_month")
        ov_year = int(colov2.number_input("Year (Overview/Calendar)",
                                          min_value=2020, max_value=2100,
                                          value=int(year), step=1, key="ov_year"))
        ov_ph = sg_public_holidays(int(ov_year))
        ov_fp = ROSTER_DIR / f"{int(ov_year)}-{int(ov_month):02d}.csv"
        if ov_fp.exists():
            ov_roster = pd.read_csv(ov_fp, dtype=str)
            ov_roster["date"] = pd.to_datetime(ov_roster["date"]).dt.date
        else:
            ov_roster = build_roster(int(ov_year), int(ov_month), mode, consultants_df, defaults_df, juniors_df, ov_ph, duties_df, read_csv("paeds.csv"), read_csv("special_slots.csv"))

        # Normalise overview roster (NaN-safe)
        _normalise_roster_inplace(ov_roster)

        st.subheader("Calendar")
        st.markdown(calendar_markdown(ov_roster, ov_ph), unsafe_allow_html=True)

        st.subheader("Overview list")
        ov_list = overview_list(ov_roster)

        # Filter by consultant
        all_people = sorted(set(consultants_df["name"].tolist()))
        who = st.multiselect("Filter by consultant(s)", options=all_people, default=[], key="ov_filter_who")
        if who:
            mask = pd.Series(False, index=ov_list.index)
            cols_to_search = [c for c in ov_list.columns if any(x in c for x in ["Room18_","Room28_","Preceptor_","Room29_"])]
            for person in who:
                m = ov_list[cols_to_search].apply(lambda s: s.astype(str).str.contains(fr"\b{person}\b", na=False))
                mask = mask | m.any(axis=1)
            ov_list = ov_list[mask]
        st.dataframe(ov_list, use_container_width=True, height=480)

    # Exports
    with tabs[-1]:
        st.subheader("Exports")
        xmonth, xyear = int(ov_month), int(ov_year)  # export the viewed month
        x_fp = ROSTER_DIR / f"{xyear}-{xmonth:02d}.csv"
        if x_fp.exists():
            xdf = pd.read_csv(x_fp, dtype=str); xdf["date"] = pd.to_datetime(xdf["date"]).dt.date
        else:
            xdf = ov_roster

        # Normalise export dataframe (NaN-safe)
        _normalise_roster_inplace(xdf)

        st.download_button("Download Roster CSV", data=xdf.to_csv(index=False).encode("utf-8"), file_name=f"Roster_{xyear}-{xmonth:02d}.csv", mime="text/csv")
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        wb = Workbook(); ws1 = wb.active; ws1.title="Roster"
        for r in dataframe_to_rows(xdf, index=False, header=True): ws1.append(r)
        ws2 = wb.create_sheet("Overview_List")
        ov = overview_list(xdf)
        for r in dataframe_to_rows(ov, index=False, header=True): ws2.append(r)
        ws3 = wb.create_sheet("Export_Flat")
        for r in dataframe_to_rows(xdf, index=False, header=True): ws3.append(r)
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        st.download_button("Download Excel Workbook", data=bio.getvalue(), file_name=f"NTBSC_Roster_{xyear}-{xmonth:02d}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.caption("Workbook contains: Roster, Overview_List, Export_Flat")

if __name__ == "__main__":
    main()
